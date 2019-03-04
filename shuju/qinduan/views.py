#-*- coding: utf-8 -*-
from django.shortcuts import render,redirect,reverse
from django.http import HttpResponse
from django.db import connection
from django.views.generic import View
import json
from io import StringIO,BytesIO
import openpyxl
from datetime import *
import calendar
from django.template import defaultfilters,defaulttags
def get_corsor():
    return connection.cursor()
def index(request):
    return render(request,'index.html')
def diaoyu(request):
    if request.method == 'GET':
        return render(request, 'diaoyu.html')
    elif request.method == 'POST':
        global d_kasj, d_jssj
        d_kasj = request.POST.get('kssj')
        d_jssj = request.POST.get('jssj')
        return render(request, 'diaoyuxq.html')
def diaoyuxq(request):
    draw = int(request.POST.get('draw',0))  # 記錄操作次數
    start = int(request.POST.get('start',0))  # 起始位置
    length = int(request.POST.get('length',0))  # 每頁長度
    data = d_select_by_page(start, start + length)
    count = d_all_count()
    dic = {
        'draw': draw,
        'recordsTotal': count,
        'recordsFiltered': count,
        'data': data,
    }
    return HttpResponse(json.dumps(dic), content_type='application/json')
def d_select_by_page(start, end):
    cursor = connection.cursor()
    sql = """
        select * from
            (select distinct dbname,db_number,accountname,rolename,roleid,bang_id,get_count,tag,tongbao,time,to_char(tbcost),biaoji,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级,rownum as rn
            from diaoyu where stat_date<date'{0}' and stat_date>=date'{1}' order by dbname )
        where rn>=%s and rn<%s
        """.format(d_jssj,d_kasj)
    cursor.execute(sql, [start, end])
    result = cursor.fetchall()
    cursor.close()
    return result

def d_all_count():
    cursor = connection.cursor()
    sql = "select count(*) from diaoyu where stat_date<date'{0}' and stat_date>=date'{1}'".format(d_jssj,d_kasj)
    cursor.execute(sql)
    result = cursor.fetchone()
    cursor.close()
    return result[0]
def d_daochu(request):
    cursor = connection.cursor()
    sql = '''select distinct * from diaoyu  where stat_date<date'{0}' and stat_date>=date'{1}' '''.format(d_jssj,d_kasj)
    cursor.execute(sql)
    deta = cursor.fetchall()
    title =[i[0] for i in cursor.description]
    cursor.close()
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename={}钓鱼数据.xlsx'.format(d_kasj)
    wb=openpyxl.Workbook()
    ws=wb.active
    for col in range(len(title)):
        c=col + 1
        ws.cell(row=1,column=c).value=title[col]
    for row in range(len(deta)):
        ws.append(deta[row])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    return response
def huizong(request):
    if request.method == 'GET':
        return render(request, 'huizong.html')
    elif request.method == 'POST':
        global h_kasj,h_jssj
        h_kasj = request.POST.get('kssj')
        h_jssj = request.POST.get('jssj')
        # sz_sj = datetime.strptime(kasj, '%Y-%m-%d')
        # od = timedelta(days=1)
        # if sz_sj.weekday() != calendar.SUNDAY:
        #     while sz_sj.weekday() != calendar.SUNDAY:
        #         sz_sj -= od
        # else:
        #     sz_sj = sz_sj - timedelta(days=7)
        # sz = sz_sj.strftime('%Y-%m-%d')
        # cur = get_corsor()
        # cur.callproc('p_renwu',[kasj, jssj, sz])
        # cur.execute('select * from huizong')
        return render(request, 'huizongxq.html')
def huizongxq(request):
    draw = int(request.POST.get('draw',0))  # 記錄操作次數
    start = int(request.POST.get('start',0))  # 起始位置
    length = int(request.POST.get('length',0))  # 每頁長度
    data = h_select_by_page(start, start + length)
    count = h_all_count()
    dic = {
        'draw': draw,
        'recordsTotal': count,
        'recordsFiltered': count,
        'data': data,
    }
    return HttpResponse(json.dumps(dic), content_type='application/json')
# def f_select_all():
#     cursor = connection.cursor()
#     sql = '''select d_bname,accountname,dbname,rolename,renwushu,roleid,tag,tongbao,time,biaoji,role_type,
#     dltag,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级 from renwu where record_date<date'{0}' and record_date>=date'{1}' order by dbname'''.format(jssj,kasj)
#     cursor.execute(sql)
#     result = cursor.fetchall()
#     cursor.close()
#     return result

def h_select_by_page(start, end):
    cursor = connection.cursor()
    sql = """
        select * from
            (select d_bname,accountname,rolename,roleid,mc,bangname,rw,tag,tongbao,time,to_char(tbcost),biaoji,to_char(alltime),role_type,dltag,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级,rownum as rn
            from huizong where record_date<date'{0}' and record_date>=date'{1}' order by dbname )
        where rn>=%s and rn<%s
        """.format(h_jssj,h_kasj)
    cursor.execute(sql, [start, end])
    result = cursor.fetchall()
    cursor.close()
    return result

def h_all_count():
    cursor = connection.cursor()
    sql = "select count(*) from huizong where record_date<date'{0}' and record_date>=date'{1}'".format(h_jssj,h_kasj)
    cursor.execute(sql)
    result = cursor.fetchone()
    cursor.close()
    return result[0]
def h_daochu(request):
    cursor = connection.cursor()
    sql = '''select distinct * from huizong  where record_date<date'{0}' and record_date>=date'{1}' '''.format(h_jssj,h_kasj)
    cursor.execute(sql)
    deta = cursor.fetchall()
    title =[i[0] for i in cursor.description]
    cursor.close()
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename={}汇总数据.xlsx'.format(h_kasj)
    wb=openpyxl.Workbook()
    ws=wb.active
    for col in range(len(title)):
        c=col + 1
        ws.cell(row=1,column=c).value=title[col]
    for row in range(len(deta)):
        ws.append(deta[row])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    return response
def renwu(request):
    if request.method == 'GET':
        return render(request, 'renwu.html')
    elif request.method == 'POST':
        global kasj,jssj
        kasj = request.POST.get('kssj')
        jssj = request.POST.get('jssj')
        return render(request, 'renwuxq.html')
def renwuxq(request):
    draw = int(request.POST.get('draw',0))  # 記錄操作次數
    start = int(request.POST.get('start',0))  # 起始位置
    length = int(request.POST.get('length',0))  # 每頁長度
    search_key = request.POST.get('search[value]',0)  # 搜索關鍵字
    order_column = request.POST.get('order[0][column]')  # 排序字段索引
    order_column = request.POST.get('order[0][dir]')  #排序規則：ase/desc
    if search_key:
        result, count = query(search_key)
        data = result[start:start + length]
    else:
        data = select_by_page(start, start + length)
        count = all_count()

    dic = {
        'draw': draw,
        'recordsTotal': count,
        'recordsFiltered': count,
        'data': data,
    }
    return HttpResponse(json.dumps(dic), content_type='application/json')
def select_all():
    cursor = connection.cursor()
    sql = '''select distinct d_bname,accountname,dbname,rolename,renwushu,roleid,tag,tongbao,time,to_char(tbcost),biaoji,to_char(alltime),role_type,
    dltag,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级 from renwu where record_date<date'{0}' and record_date>=date'{1}' order by dbname'''.format(jssj,kasj)
    cursor.execute(sql)
    result = cursor.fetchall()
    cursor.close()
    return result

def select_by_page(start, end):
    cursor = connection.cursor()
    sql = """
        select * from
            (select distinct d_bname,accountname,dbname,rolename,renwushu,roleid,tag,tongbao,time,to_char(tbcost),biaoji,to_char(alltime),role_type,dltag,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级,rownum as rn
            from renwu where record_date<date'{0}' and record_date>=date'{1}' order by dbname )
        where rn>=%s and rn<%s
        """.format(jssj,kasj)
    cursor.execute(sql, [start, end])
    result = cursor.fetchall()
    cursor.close()
    return result

def all_count():
    cursor = connection.cursor()
    sql = "select count(*) from renwu where record_date<date'{0}' and record_date>=date'{1}'".format(jssj,kasj)
    cursor.execute(sql)
    result = cursor.fetchone()
    cursor.close()
    return result[0]

def query(search_key):
    cursor = connection.cursor()
    sql = '''select distinct d_bname,accountname,dbname,rolename,renwushu,roleid,tag,tongbao,time,to_char(tbcost),biaoji,to_char(alltime),role_type,
    dltag,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级 from renwu where dbname like '%%{0}%%' and record_date<date'{1}' and record_date>=date'{2}' '''.format(search_key,jssj,kasj)
    cursor.execute(sql)
    result = cursor.fetchall()
    sql = '''select count(*) from renwu where dbname like '%%{0}%%' and record_date<date'{1}' and record_date>=date'{2}' '''.format(search_key,jssj,kasj)
    cursor.execute(sql)
    count = cursor.fetchone()[0]
    cursor.close()
    return result, count
def quanliang(request):
    cursor = connection.cursor()
    sql = '''select distinct d_bname,accountname,dbname,rolename,renwushu,roleid,tag,tongbao,time,tbcost,biaoji,alltime,role_type,dltag,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级 
    from renwu where record_date<date'{0}' and record_date>=date'{1}'order by dbname'''.format(jssj,kasj)
    cursor.execute(sql)
    deta = cursor.fetchall()
    title =[i[0] for i in cursor.description]
    cursor.close()
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename={}任务数据.xlsx'.format(kasj)
    wb=openpyxl.Workbook()
    ws=wb.active
    for col in range(len(title)):
        c=col + 1
        ws.cell(row=1,column=c).value=title[col]
    for row in range(len(deta)):
        ws.append(deta[row])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    return response

def fuben(request):
    if request.method=='GET':
        return render(request,'fuben.html')
    elif request.method=='POST':
        global f_jssj,f_kasj
        f_kasj=request.POST.get('kssj')
        f_jssj=request.POST.get('jssj')
        return render(request, 'fubenxq.html')
def fubenxq(request):
    draw = int(request.POST.get('draw',0))  # 記錄操作次數
    start = int(request.POST.get('start',0))  # 起始位置
    length = int(request.POST.get('length',0))  # 每頁長度
    search_key = request.POST.get('search[value]',0)  # 搜索關鍵字
    order_column = request.POST.get('order[0][column]')  # 排序字段索引
    order_column = request.POST.get('order[0][dir]')  #排序規則：ase/desc
    if search_key:
        result, count = f_query(search_key)
        data = result[start:start + length]
    else:
        data = f_select_by_page(start, start + length)
        count = f_all_count()

    dic = {
        'draw': draw,
        'recordsTotal': count,
        'recordsFiltered': count,
        'data': data,
    }
    return HttpResponse(json.dumps(dic), content_type='application/json')
# def f_select_all():
#     cursor = connection.cursor()
#     sql = '''select d_bname,accountname,dbname,rolename,renwushu,roleid,tag,tongbao,time,biaoji,role_type,
#     dltag,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级 from renwu where record_date<date'{0}' and record_date>=date'{1}' order by dbname'''.format(jssj,kasj)
#     cursor.execute(sql)
#     result = cursor.fetchall()
#     cursor.close()
#     return result

def f_select_by_page(start, end):
    cursor = connection.cursor()
    sql = """
        select * from
            (select distinct d_bname,dbname,accountname,rolename,roleid,count(*),tag,tongbao,time,to_char(tbcost),biaoji,to_char(alltime),role_type,dltag,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级,map_name,rownum as rn
            from fuben where record_date<date'{0}' and record_date>=date'{1}' group by d_bname,dbname,accountname,rolename,roleid,tag,tongbao,time,tbcost,biaoji,alltime,role_type,dltag,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级,map_name,rownum order by dbname )
        where rn>=%s and rn<%s
        """.format(f_jssj,f_kasj)
    cursor.execute(sql, [start, end])
    result = cursor.fetchall()
    cursor.close()
    return result

def f_all_count():
    cursor = connection.cursor()
    sql = "select count(*) from fuben where record_date<date'{0}' and record_date>=date'{1}'".format(f_jssj,f_kasj)
    cursor.execute(sql)
    result = cursor.fetchone()
    cursor.close()
    return result[0]

def f_query(search_key):
    cursor = connection.cursor()
    sql = '''select distinct d_bname,dbname,accountname,rolename,roleid,count(*),tag,tongbao,time,to_char(tbcost),biaoji,to_char(alltime),role_type,dltag,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级,map_name
     from fuben where dbname like '%%{0}%%' and record_date<date'{1}' and record_date>=date'{2}' group by d_bname,dbname,accountname,rolename,roleid,tag,tongbao,time,tbcost,biaoji,alltime,role_type,dltag,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级,map_name '''.format(search_key,f_jssj,f_kasj)
    cursor.execute(sql)
    result = cursor.fetchall()
    sql = '''select count(*) from fuben where dbname like '%%{0}%%' and record_date<date'{1}' and record_date>=date'{2}' '''.format(search_key,f_jssj,f_kasj)
    cursor.execute(sql)
    count = cursor.fetchone()[0]
    cursor.close()
    return result, count
def f_quanliang(request):
    cursor = connection.cursor()
    sql = '''select distinct d_bname,dbname,accountname,rolename,roleid,count(*),tag,tongbao,time,tbcost,biaoji,alltime,role_type,dltag,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级,map_name 
    from fuben where record_date<date'{0}' and record_date>=date'{1}' group by d_bname,dbname,accountname,rolename,roleid,tag,tongbao,time,tbcost,biaoji,alltime,role_type,dltag,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级,map_name order by dbname '''.format(f_jssj,f_kasj)
    cursor.execute(sql)
    deta = cursor.fetchall()
    title =[i[0] for i in cursor.description]
    cursor.close()
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename={}副本数据.xlsx'.format(f_kasj)
    wb=openpyxl.Workbook()
    ws=wb.active
    for col in range(len(title)):
        c=col + 1
        ws.cell(row=1,column=c).value=title[col]
    for row in range(len(deta)):
        ws.append(deta[row])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    return response
class dengluxx(View):
    def get(self,request):
        return render(request,'denglu.html')
    def post(self,request):
        self.uu=request.POST.get('uu')
        self.kasj=request.POST.get('kssj')
        self.jssj=request.POST.get('jssj')
        cur = get_corsor()
        cur.execute('''select regexp_substr(b.description,'[^-()]+',1,2) dbname,a.* from
        (select /*+parallel(r,2)+*/accountname,dbname,rolename,roleid,col_3,col_6,count(*) ccs FROM xsjods.jx3_role_log_opeartion r
        WHERE record_date>=date'{0}'
        AND record_date<date'{1}'
        AND actiontype=139
        and col_3='{2}'
        group by accountname,dbname,rolename,roleid,col_3,col_6)a
        inner join
        xsjdw.v_dim_jx3_server  b
        on substr(a.dbname,-4,4)=b.db_number'''.format(self.kasj, self.jssj, self.uu))
        self.deta = cur.fetchall()
        return render(request, 'dengluxq.html', context={'deta': self.deta})

    # def output(self,request):
    #     response = HttpResponse(content_type='application/vnd.ms-excel')
    #     response['Content-Disposition'] = 'attachment;filename={}登录信息.xlsx'.format(self.jssj)
    #     wb=openpyxl.Workbook()
    #     ws=wb.active
    #     for col in range(len(self.title)):
    #         c=col + 1
    #         ws.cell(row=1,column=c).value=self.title[col]
    #     for row in range(len(self.deta)):
    #         ws.append(self.deta[row])
    #     output = BytesIO()
    #     wb.save(output)
    #     output.seek(0)
    #     response.write(output.getvalue())
    #     return response


    # response = HttpResponse(content_type='application/vnd.ms-excel')
    # response['Content-Disposition'] = 'attachment;filename=user.xls'
    # wb = xlwt.Workbook(encoding = 'utf-8')
    # sheet = wb.add_sheet(u'人员表单')
    # for t in range(len(title)):
    #     for i in title:
    #         sheet.write(0,t,i)
    # output = BytesIO()
    # wb.save(output)
    # output.seek(0)
    # response.write(output.getvalue())
    # return response
    # def put(self,request):
    #     response = HttpResponse(content_type='application/vnd.ms-excel')
    #     response['Content-Disposition'] = 'attachment;filename={}登录数据.xls'.format(self.jssj)
    #     fm = pd.DataFrame(self.deta, columns=self.title)
    #     fm.to_excel('登录数据.xlsx'.format(self.jssj), index=False, encoding='utf-8')
    #     output = BytesIO()
    #     fm.save(output)
    #     output.seek(0)
    #     response.write(output.getvalue())
    #     return response
# class denglu(View):
#     def get(self,request):
#         form = mesesagecanshu()
#         return render(request,'denglu.html',context={'form':form})
#     def post(self,request):
#         form = mesesagecanshu(request.POST)
#         if form.is_valid():
#             uu = form.cleaned_data.get('UU')
#             kajs = form.cleaned_data.get('kssj')
#             jssj = form.cleaned_data.get('jssj')
#             #conte = {'uu':uu,'kssj':kajs,'jssj':jssj}
#             #return render(request,'dengluxq.html',context={'deta':deta,'title':title})
#             #return redirect('dengluxq/{}/{}/{}/'.format(uu,kajs,jssj))
#             cur = get_corsor()
#             cur.execute('''select regexp_substr(b.description,'[^-()]+',1,2) dbname,a.* from
#             (select /*+parallel(r,2)+*/accountname,dbname,rolename,roleid,col_3,col_6,count(*) ccs FROM xsjods.jx3_role_log_opeartion r
#             WHERE record_date>=date'{0}'
#             AND record_date<date'{1}'
#             AND actiontype=139
#             and col_3='{2}'
#             group by accountname,dbname,rolename,roleid,col_3,col_6)a
#             inner join
#             xsjdw.v_dim_jx3_server  b
#             on substr(a.dbname,-4,4)=b.db_number'''.format(kajs, jssj, uu))
#             deta = cur.fetchall()[0:50]
#             print(deta)
#             title = [i[0] for i in cur.description]
#             return render(request, 'dengluxq.html', context={'deta': deta, 'title': title})
#         else:
#             print(form.errors.get_json_data())
#             return HttpResponse('错误')
# def dengluxq(request,uu,kajs,jssj):
#     pass
#     #uu,kajs,jssj=request.GET.get('id')
#     # cur = get_corsor()
#     # cur.execute('''select regexp_substr(b.description,'[^-()]+',1,2) dbname,a.* from
#     # (select /*+parallel(r,2)+*/accountname,dbname,rolename,roleid,col_3,col_6,count(*) ccs FROM xsjods.jx3_role_log_opeartion r
#     # WHERE record_date>=date'{0}'
#     # AND record_date<date'{1}'
#     # AND actiontype=139
#     # and col_3='{2}'
#     # group by accountname,dbname,rolename,roleid,col_3,col_6)a
#     # inner join
#     # xsjdw.v_dim_jx3_server  b
#     # on substr(a.dbname,-4,4)=b.db_number'''.format(kajs, jssj, uu))
#     # deta = cur.fetchall()
#     # title = [i[0] for i in cur.description]
#     # return render(request,'dengluxq.html',context={'deta':deta,'title':title})
def dangrirw(request):
    if request.method == 'GET':
        return render(request, 'dangrirw.html')
    elif request.method == 'POST':
        global dr_kasj, dr_jssj
        dr_kasj = request.POST.get('kasj')
        dr_jssj = request.POST.get('jssj')
        now = datetime.now().date()
        djssj = now.strftime('%Y-%m-%d')
        od = timedelta(days=1)
        if now.weekday() != calendar.SUNDAY:
            while now.weekday() != calendar.SUNDAY:
                now -= od
        else:
            now = now - timedelta(days=7)
        szsj = now.strftime('%Y-%m-%d')
        cursor = connection.cursor()
        # cursor.execute('select min(record_date) min_da,max(record_date) max_da from drrenwu')
        # da =cursor.fetchall()
        # print(da)
        cursor.callproc('p_dangrirw', [dr_kasj, dr_jssj, djssj, szsj])
        return render(request, 'drrwxq.html')


def drrenwuxq(request):
    draw = int(request.POST.get('draw', 0))  # 記錄操作次數
    start = int(request.POST.get('start', 0))  # 起始位置
    length = int(request.POST.get('length', 0))  # 每頁長度
    search_key = request.POST.get('search[value]', 0)  # 搜索關鍵字
    order_column = request.POST.get('order[0][column]')  # 排序字段索引
    order_column = request.POST.get('order[0][dir]')  # 排序規則：ase/desc

    data = r_select_by_page(start, start + length)
    count = r_all_count()

    dic = {
        'draw': draw,
        'recordsTotal': count,
        'recordsFiltered': count,
        'data': data,
    }
    return HttpResponse(json.dumps(dic), content_type='application/json')


# def r_select_all():
#     cursor = connection.cursor()
#     sql = '''select d_bname,accountname,dbname,rolename,renwushu,roleid,tag,tongbao,time,biaoji,role_type,
#         dltag,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级 from renwu where record_date<date'{0}' and record_date>=date'{1}' order by dbname'''.format(
#         jssj, kasj)
#     cursor.execute(sql)
#     result = cursor.fetchall()
#     cursor.close()
#     return result


def r_select_by_page(start, end):
    # print(dr_kasj)
    # dr_ka=str(dr_kasj).split('')[0]
    # dr_js=str(dr_jssj).split('')[0]
    cursor = connection.cursor()
    sql = """select * from
(select d_bname,accountname,dbname,rolename,renwushu,roleid,tag,tongbao,time,to_char(tbcost) 通宝消耗,装备分数,最大精练,最大镶嵌,精练等级,镶嵌等级,rownum as rn from drrenwu )
where rn>=%s and rn<%s
""".format(dr_kasj,dr_jssj)
    cursor.execute(sql, [start, end])
    result = cursor.fetchall()
    cursor.close()
    return result


def r_all_count():
    cursor = connection.cursor()
    sql = '''select count(*) from drrenwu'''
    cursor.execute(sql)
    result = cursor.fetchone()
    cursor.close()
    return result[0]


def r_query(search_key):
    cursor = connection.cursor()
    sql = '''select * from drrenwu where dbname like '%%{0}%%' '''.format(
        search_key)
    cursor.execute(sql)
    result = cursor.fetchall()
    sql = '''select count(*) from drrenwu where dbname like '%%{0}%%' '''.format(
        search_key)
    cursor.execute(sql)
    count = cursor.fetchone()[0]
    cursor.close()
    return result, count


def r_quanliang(request):
    cursor = connection.cursor()
    sql = '''select * from drrenwu'''
    cursor.execute(sql)
    deta = cursor.fetchall()
    title = [i[0] for i in cursor.description]
    cursor.close()
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename={}任务数据.xlsx'.format(dr_kasj)
    wb = openpyxl.Workbook()
    ws = wb.active
    for col in range(len(title)):
        c = col + 1
        ws.cell(row=1, column=c).value = title[col]
    for row in range(len(deta)):
        ws.append(deta[row])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    return response